"""
Question bank service.
"""
from __future__ import annotations

import csv
import io
import importlib
import json
import os
import shutil
import time
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

from fastapi import HTTPException, UploadFile, status
from sqlalchemy.orm import Session

from repositories.question_bank_repo import QuestionBankRepository, normalize_stem


QUESTION_BANK_UPLOAD_DIR = os.path.join("uploads", "question-bank")
os.makedirs(QUESTION_BANK_UPLOAD_DIR, exist_ok=True)

IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp", ".gif"}
IMPORT_EXTENSIONS = {".csv", ".xlsx", ".xls"}


class QuestionBankService:
    @staticmethod
    def _normalize_asset_file_path(file_path: Optional[str]) -> str:
        normalized = str(file_path or "").replace("\\", "/").strip()
        normalized = normalized.lstrip("/")
        if normalized.startswith("./"):
            normalized = normalized[2:]
        return normalized

    @staticmethod
    def _normalize_expected_timestamp(value: Optional[str]) -> Optional[str]:
        text = str(value or "").strip()
        return text or None

    @staticmethod
    def _merge_metadata(
        existing: Optional[Dict[str, Any]],
        incoming: Optional[Dict[str, Any]],
    ) -> Dict[str, Any]:
        merged: Dict[str, Any] = {}
        if isinstance(existing, dict):
            merged.update(existing)
        if isinstance(incoming, dict):
            merged.update(incoming)
        return merged

    @classmethod
    def _build_editor_metadata(
        cls,
        *,
        existing: Optional[Dict[str, Any]] = None,
        incoming: Optional[Dict[str, Any]] = None,
        created_by: Optional[int] = None,
        created_by_name: Optional[str] = None,
        edited_by: Optional[int] = None,
        edited_by_name: Optional[str] = None,
    ) -> Dict[str, Any]:
        metadata = cls._merge_metadata(existing, incoming)
        now = datetime.utcnow().isoformat()

        if created_by is not None and metadata.get("created_by_id") is None:
            metadata["created_by_id"] = created_by
        if created_by_name and not metadata.get("created_by_name"):
            metadata["created_by_name"] = created_by_name
        if metadata.get("created_by_id") is None and edited_by is not None:
            metadata["created_by_id"] = edited_by
        if not metadata.get("created_by_name") and edited_by_name:
            metadata["created_by_name"] = edited_by_name

        if edited_by is not None:
            metadata["last_edited_by_id"] = edited_by
        if edited_by_name:
            metadata["last_edited_by_name"] = edited_by_name
        if edited_by is not None or edited_by_name:
            metadata["last_edited_at"] = now

        return metadata

    @staticmethod
    def _assert_edit_version(item: Any, expected_updated_at: Optional[str]) -> None:
        expected = QuestionBankService._normalize_expected_timestamp(expected_updated_at)
        if not expected:
            return
        actual = item.updated_at.isoformat() if getattr(item, "updated_at", None) else None
        if actual != expected:
            raise ValueError("question bank item has been updated by another admin")

    @staticmethod
    def _read_excel_rows(raw_bytes: bytes) -> List[Dict[str, Any]]:
        openpyxl_spec = importlib.util.find_spec("openpyxl")
        if openpyxl_spec is not None:
            from openpyxl import load_workbook

            workbook = load_workbook(io.BytesIO(raw_bytes), data_only=True)
            worksheet = workbook.active
            rows = list(worksheet.iter_rows(values_only=True))
            if not rows:
                return []

            headers = [
                str(cell).strip() if cell is not None else ""
                for cell in rows[0]
            ]
            normalized_headers = [
                header if header else f"column_{index + 1}"
                for index, header in enumerate(headers)
            ]

            parsed_rows: List[Dict[str, Any]] = []
            for row_values in rows[1:]:
                row_dict: Dict[str, Any] = {}
                for index, header in enumerate(normalized_headers):
                    value = row_values[index] if index < len(row_values) else ""
                    row_dict[header] = "" if value is None else value
                if any(str(value).strip() for value in row_dict.values()):
                    parsed_rows.append(row_dict)
            return parsed_rows

        pandas_spec = importlib.util.find_spec("pandas")
        if pandas_spec is not None:
            pandas = importlib.import_module("pandas")
            dataframe = pandas.read_excel(io.BytesIO(raw_bytes))
            return dataframe.fillna("").to_dict(orient="records")

        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Excel import requires openpyxl or pandas to be installed",
        )

    @staticmethod
    def _as_list(value: Any) -> List[str]:
        if value is None:
            return []
        if isinstance(value, list):
            return [str(item).strip() for item in value if str(item).strip()]
        if isinstance(value, str):
            return [item.strip() for item in value.replace("；", ",").replace("、", ",").split(",") if item.strip()]
        return [str(value).strip()]

    @staticmethod
    def _parse_answer(value: Any) -> Any:
        if isinstance(value, list):
            return value
        if isinstance(value, dict):
            return value
        if value is None:
            return ""
        text = str(value).strip()
        if not text:
            return ""
        if text.startswith("[") and text.endswith("]"):
            try:
                parsed = json.loads(text)
                if isinstance(parsed, list):
                    return parsed
            except Exception:  # pylint: disable=broad-except
                return text
        return text

    @staticmethod
    def _parse_options(value: Any) -> List[str]:
        if value is None:
            return []
        if isinstance(value, list):
            return [str(item).strip() for item in value if str(item).strip()]
        text = str(value).strip()
        if not text:
            return []
        if text.startswith("[") and text.endswith("]"):
            try:
                parsed = json.loads(text)
                if isinstance(parsed, list):
                    return [str(item).strip() for item in parsed if str(item).strip()]
            except Exception:  # pylint: disable=broad-except
                pass
        return [item.strip() for item in text.replace("；", "\n").replace("||", "\n").splitlines() if item.strip()]

    @staticmethod
    def _serialize_asset(asset) -> Dict[str, Any]:
        normalized_path = QuestionBankService._normalize_asset_file_path(asset.file_path)
        preview_path = normalized_path if normalized_path.startswith("uploads/") else f"uploads/{normalized_path}"
        preview_url = f"/{preview_path}"
        return {
            "id": asset.id,
            "asset_type": asset.asset_type,
            "file_name": asset.file_name,
            "file_path": normalized_path,
            "mime_type": asset.mime_type,
            "sort_order": asset.sort_order,
            "preview_url": preview_url,
        }

    @classmethod
    def serialize_item(cls, item) -> Dict[str, Any]:
        assets = [cls._serialize_asset(asset) for asset in (item.assets or [])]
        grouped = {
            "question_images": [asset for asset in assets if asset["asset_type"] == "question_image"],
            "answer_images": [asset for asset in assets if asset["asset_type"] == "answer_image"],
            "solution_images": [asset for asset in assets if asset["asset_type"] == "solution_image"],
        }
        return {
            "id": item.id,
            "stem": item.stem,
            "question_type": item.question_type,
            "grade_level": item.grade_level,
            "subject": item.subject,
            "difficulty": item.difficulty,
            "knowledge_points": item.knowledge_points or [],
            "answer": item.answer,
            "explanation": item.explanation,
            "options": item.options or [],
            "source": item.source,
            "source_type": item.source_type,
            "status": item.status,
            "created_by": item.created_by,
            "metadata": item.metadata_json or {},
            "assets": assets,
            **grouped,
            "created_at": item.created_at.isoformat() if item.created_at else None,
            "updated_at": item.updated_at.isoformat() if item.updated_at else None,
        }

    @classmethod
    def validate_item_payload(cls, payload: Dict[str, Any]) -> Dict[str, Any]:
        stem = str(payload.get("stem") or "").strip()
        if not stem:
            raise ValueError("stem is required")
        question_type = str(payload.get("question_type") or "").strip()
        if not question_type:
            raise ValueError("question_type is required")
        answer = cls._parse_answer(payload.get("answer"))
        if answer in ("", [], None):
            raise ValueError("answer is required")
        return {
            "stem": stem,
            "normalized_stem": normalize_stem(stem),
            "question_type": question_type,
            "grade_level": cls._normalize_grade_level(payload),
            "subject": cls._normalize_subject(payload),
            "difficulty": str(payload.get("difficulty") or "").strip() or None,
            "knowledge_points": cls._as_list(payload.get("knowledge_points")),
            "answer": answer,
            "explanation": str(payload.get("explanation") or "").strip() or None,
            "options": cls._parse_options(payload.get("options")),
            "source": str(payload.get("source") or "").strip() or None,
            "source_type": str(payload.get("source_type") or "question_bank").strip(),
            "status": str(payload.get("status") or "active").strip(),
            "created_by": payload.get("created_by"),
            "metadata_json": payload.get("metadata") or {},
            "expected_updated_at": cls._normalize_expected_timestamp(payload.get("expected_updated_at")),
        }

    @staticmethod
    def _normalize_grade_level(payload: Dict[str, Any]) -> Optional[str]:
        from utils.question_bank_constants import normalize_grade_level as _norm_gl
        raw = str(payload.get("grade_level") or "").strip() or None
        if raw is None:
            return None
        normalized = _norm_gl(raw)
        if normalized is None:
            raise ValueError(f"invalid grade_level: '{raw}'")
        return normalized

    @staticmethod
    def _normalize_subject(payload: Dict[str, Any]) -> Optional[str]:
        from utils.question_bank_constants import normalize_subject as _norm_subj
        raw = str(payload.get("subject") or "").strip() or None
        if raw is None:
            return None
        normalized = _norm_subj(raw)
        if normalized is None:
            raise ValueError(f"invalid subject: '{raw}'")
        return normalized

    @classmethod
    def list_items(cls, db: Session, **filters: Any) -> Dict[str, Any]:
        items = QuestionBankRepository.list_items(db, **filters)
        total = QuestionBankRepository.count_items(
            db,
            grade_level=filters.get("grade_level"),
            subject=filters.get("subject"),
            difficulty=filters.get("difficulty"),
            question_type=filters.get("question_type"),
            status=filters.get("status"),
            keyword=filters.get("keyword"),
        )
        return {
            "items": [cls.serialize_item(item) for item in items],
            "total": total,
        }

    @classmethod
    def create_item(
        cls,
        db: Session,
        payload: Dict[str, Any],
        created_by: Optional[int] = None,
        created_by_name: Optional[str] = None,
    ) -> Dict[str, Any]:
        data = cls.validate_item_payload({**payload, "created_by": created_by})
        metadata = cls._build_editor_metadata(
            incoming=data.get("metadata_json"),
            created_by=created_by,
            created_by_name=created_by_name,
            edited_by=created_by,
            edited_by_name=created_by_name,
        )
        data["metadata_json"] = metadata
        data.pop("expected_updated_at", None)
        duplicate = QuestionBankRepository.find_by_normalized_stem(
            db,
            data["normalized_stem"],
            grade_level=data.get("grade_level"),
            subject=data.get("subject"),
        )
        if duplicate:
            raise ValueError("duplicate question detected")
        item = QuestionBankRepository.create_item(db, **data)
        return cls.serialize_item(QuestionBankRepository.get_item(db, item.id))

    @classmethod
    def update_item(
        cls,
        db: Session,
        item_id: int,
        payload: Dict[str, Any],
        updated_by: Optional[int] = None,
        updated_by_name: Optional[str] = None,
    ) -> Dict[str, Any]:
        data = cls.validate_item_payload(payload)
        current_item = QuestionBankRepository.get_item(db, item_id)
        if not current_item:
            raise ValueError("question bank item not found")
        cls._assert_edit_version(current_item, data.get("expected_updated_at"))
        data["metadata_json"] = cls._build_editor_metadata(
            existing=current_item.metadata_json,
            incoming=data.get("metadata_json"),
            created_by=current_item.created_by,
            edited_by=updated_by,
            edited_by_name=updated_by_name,
        )
        data.pop("expected_updated_at", None)
        duplicate = QuestionBankRepository.find_by_normalized_stem(
            db,
            data["normalized_stem"],
            grade_level=data.get("grade_level"),
            subject=data.get("subject"),
        )
        if duplicate and duplicate.id != item_id:
            raise ValueError("duplicate question detected")
        item = QuestionBankRepository.update_item(db, item_id, **data)
        return cls.serialize_item(QuestionBankRepository.get_item(db, item.id))

    @staticmethod
    def delete_item(db: Session, item_id: int) -> bool:
        return QuestionBankRepository.delete_item(db, item_id)

    @staticmethod
    def _safe_filename(file_name: str) -> str:
        name, ext = os.path.splitext(file_name or "upload")
        safe_name = "".join(ch if ch.isalnum() or ch in {"-", "_"} else "_" for ch in name)[:80] or "upload"
        return f"{safe_name}_{int(time.time() * 1000)}{ext.lower()}"

    @classmethod
    async def save_asset(
        cls,
        *,
        file: UploadFile,
        item_id: int,
        asset_type: str,
        sort_order: int = 0,
        db: Session,
    ) -> Dict[str, Any]:
        _, ext = os.path.splitext(file.filename or "")
        if ext.lower() not in IMAGE_EXTENSIONS:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Only image assets are supported")
        relative_dir = os.path.join("question-bank", str(item_id))
        absolute_dir = os.path.abspath(os.path.join("uploads", relative_dir))
        os.makedirs(absolute_dir, exist_ok=True)
        saved_name = cls._safe_filename(file.filename or "image.png")
        absolute_path = os.path.join(absolute_dir, saved_name)
        with open(absolute_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        relative_path = cls._normalize_asset_file_path(os.path.join(relative_dir, saved_name))
        asset = QuestionBankRepository.create_asset(
            db,
            item_id=item_id,
            asset_type=asset_type,
            file_name=file.filename or saved_name,
            file_path=relative_path,
            mime_type=file.content_type,
            sort_order=sort_order,
        )
        return cls._serialize_asset(asset)

    @classmethod
    async def parse_import_file(cls, file: UploadFile) -> List[Dict[str, Any]]:
        _, ext = os.path.splitext(file.filename or "")
        ext = ext.lower()
        if ext not in IMPORT_EXTENSIONS:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Only CSV and Excel files are supported")

        raw_bytes = await file.read()
        if ext == ".csv":
            decoded = raw_bytes.decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(decoded))
            rows = [dict(row) for row in reader]
        else:
            rows = cls._read_excel_rows(raw_bytes)
        return rows

    @classmethod
    async def import_items(
        cls,
        db: Session,
        *,
        file: UploadFile,
        created_by: Optional[int] = None,
        created_by_name: Optional[str] = None,
    ) -> Dict[str, Any]:
        rows = await cls.parse_import_file(file)
        created: List[Dict[str, Any]] = []
        duplicates = 0
        errors: List[Dict[str, Any]] = []
        for index, row in enumerate(rows, start=1):
            try:
                payload = cls.validate_item_payload({**row, "created_by": created_by})
                payload["metadata_json"] = cls._build_editor_metadata(
                    incoming=payload.get("metadata_json"),
                    created_by=created_by,
                    created_by_name=created_by_name,
                    edited_by=created_by,
                    edited_by_name=created_by_name,
                )
                payload.pop("expected_updated_at", None)
                duplicate = QuestionBankRepository.find_by_normalized_stem(
                    db,
                    payload["normalized_stem"],
                    grade_level=payload.get("grade_level"),
                    subject=payload.get("subject"),
                )
                if duplicate:
                    duplicates += 1
                    continue
                item = QuestionBankRepository.create_item(db, **payload)
                created.append(cls.serialize_item(QuestionBankRepository.get_item(db, item.id)))
            except Exception as exc:  # pylint: disable=broad-except
                errors.append({"row": index, "error": str(exc)})
        return {
            "created_count": len(created),
            "duplicate_count": duplicates,
            "error_count": len(errors),
            "items": created,
            "errors": errors,
        }

    @classmethod
    def get_item_or_404(cls, db: Session, item_id: int) -> Dict[str, Any]:
        item = QuestionBankRepository.get_item(db, item_id)
        if not item:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="question bank item not found")
        return cls.serialize_item(item)
